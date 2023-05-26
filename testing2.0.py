from ppadb.client import Client
import cv2
import sys
import os
from PIL import Image
import pytesseract
import numpy as np
import time
from matplotlib import pyplot as plt
import openpyxl
from datetime import date
import tkinter as tk
from tkinter import messagebox
import keyboard
from neural_network import read_ocr
import requests
import webbrowser
from openpyxl.styles.borders import Border, Side
import time

version = "RokTracker-v7.3"
def tointcheck(element):
    try:
        return int(element)
    except ValueError:
        return element
        
def tointprint(element):
    try:
        return str(f'{int(element):,}')
    except ValueError:
        return str(element)

#Initiliaze paths and variables
today = date.today()

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe' #Change to your installation path folder.

def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)
os.system("")

##### Update Checker #####
response = requests.get("https://api.github.com/repos/nikolakis1919/RokTracker/releases/latest")
if (response.json()["name"]) != version:
    bo = tk.Tk()
    bo.withdraw()
    messagebox.showinfo("Tool is outdated", "New version is available on github repository. It is highly recommended to update the tool!")
    bo.destroy()


####### Tkinter Section ##########
#Create input gui
root=tk.Tk()

#Tkinter title
root.title('RokTracker')

#Tkinter window size
root.geometry("400x250")

#Tkinter function
def link():
    webbrowser.open_new(r"https://www.paypal.com/donate/?hosted_button_id=55G95MNYPVX72")

#Initialize Options for dropdown box
OPTIONS = []
for i in range(38):
    OPTIONS.append(50+i*25)
    
#Variables
variable = tk.StringVar(root)
variable.set('')
variable2 = tk.IntVar(root)
variable2.set(OPTIONS[0]) # default value
var1 = tk.IntVar()

#Labels
kingdom_label = tk.Label(root, text = 'Kingdom', font=('calibre',10, 'bold'))  
search_top_label = tk.Label(root, text = 'Search Amount', font=('calibre',10, 'bold'))
#Copyrights
copyright=u"\u00A9"
l1=tk.Label(root,text=copyright + ' nikolakis1919', font = ('calibre',10,'bold')) 
l2=tk.Button(root, foreground='Green', text='Donate', command=link, font = ('calibre',10,'bold'))
l3=tk.Label(root,text='Find me on discord: nikos#4469', font = ('calibre',10,'bold'))

#Input Fields
kingdom_entry = tk.Entry(root,textvariable = variable, font=('calibre',10,'normal'))
w = tk.OptionMenu(root, variable2, *OPTIONS)
resume_scan =tk.Checkbutton(root, text="Resume Scan", variable=var1, font=('calibre',10,'bold'))

def search():
    if variable.get():
        global kingdom
        kingdom = variable.get()
        global search_range
        search_range = variable2.get()
        root.destroy()
        global resume_scanning
        resume_scanning = var1.get()
        print("Scanning Started...")
    else:
        print("You need to fill Kingdom number!")
        kingdom_entry.focus_set()
        
button = tk.Button(root, text="Search", command=search)

#Positions in tkinter Grid
kingdom_label.grid(row=0,column=0)
kingdom_entry.grid(row=0,column=1)
search_top_label.grid(row=1,column=0)
w.grid(row=1,column=1)
resume_scan.grid(row=2,column=1,pady=4)
button.grid(row=3,column=1,pady=5)
l1.grid(row=4,column=1,pady=10)
l2.grid(row=8,column=1,pady=10)
l3.grid(row=5,column=1,pady=10)
root.mainloop()

#######RokTracker
#Initialize the connection to adb
adb = Client(host='localhost', port=5037)
devices = adb.devices()

if len(devices) == 0:
    print('no device attached')
    quit()

#Prolly a good idea to have only 1 device while running this
device = devices[0]

######Excel Formatting
wb = openpyxl.Workbook()
sh = wb.create_sheet(index=0,title=str(today))



#Make Head Bold
bold = openpyxl.styles.Font(bold=True ,italic=True)
for i in range(1,30):
    sh.cell(row=1, column=i).font = bold

#aligment
Aligment = openpyxl.styles.Alignment(horizontal='center')
for i in range(1,30):
    for b in range(1,1000):
        sh.cell(row=b, column=i).alignment=Aligment
#borders
bborder = Border(left=Side('thin'),right=Side('thin'),bottom=Side('thin'),top=Side('thin'))
for i in range(18,22):
    for b in range(1,9):
        sh.cell(row=b, column=i).border = bborder
#fill
fill1 = openpyxl.styles.PatternFill(fill_type='solid',start_color='757171')
fill2 = openpyxl.styles.PatternFill(fill_type='solid',start_color='D0CECE')
fill3 = openpyxl.styles.PatternFill(fill_type='solid',start_color='FFC000')

sh.cell(row=1, column=18).fill=fill3
for i in range(19,22):
    sh.cell(row=1, column=i).fill=fill1
for i in range(2,9):
    sh.cell(row=i, column=18).fill=fill1
for i in range(19,22):
    for b in range(2,9):
        sh.cell(row=b, column=i).fill=fill2
#coulnm width
coulnm = ['A','B','C','D','E','F','G','H','I','J','N','O']
coulnm2= ['K','L','M']
coulnm3= ['R','S','T','U']
for i in coulnm :
    sh.column_dimensions[i].width = 16
for i in coulnm2 :
    sh.column_dimensions[i].width = 14
for i in coulnm3 :
    sh.column_dimensions[i].width = 20
sh.column_dimensions["P"].width = 20
sh.column_dimensions["Q"].width = 5


#Initialize Excel Sheet Header
headers=['Id','Name','Power','KillPoints','Deads','T1 Kills','T2 Kills','T3 Kills',
'T4 Kills','T5 Kills' ,'Victories','Defeats','Scouts',
'Gathered','Assitance','Alliance']
for i , b in enumerate(headers) :
    sh.cell(row=1, column=i+1).value = b
headers = [str(kingdom),'TOP300','TOP600','TOP900']
for i , b in enumerate(headers):
    sh.cell(row=1, column=i+18).value= b
headers=['Power Capped',' KillPoins','T5 Kills','T4 Kills','Deads','RssGathered','RssAssistance']
for i , b in enumerate(headers):
    sh.cell(row=i+2, column=18).value=b

#change numbers format
for i in range(2,30):
    for b in range(2,1000):
        sh.cell(row=b, column=i).number_format='#,##0'

#Position for next governor to check
Y =[285, 390, 490, 590, 605]

#Resume Scan options. Refine the loop
j = 0
if resume_scanning:
    j = 4
    search_range = search_range + j
#The loop in TOP XXX Governors in kingdom - It works both for power and killpoints Rankings
#MUST have the tab opened to the 1st governor(Power or Killpoints)


##### Save button listener#####
stop = False
def onkeypress(event):
    global stop
    if event.name == '\\':
        print("Your scan will be terminated when current governor scan is over!")
        stop = True

keyboard.on_press(onkeypress)

try:
    for i in range(j,search_range):
        start_time = time.perf_counter()
        if stop:
            print('stopped')
            #print("Scan Terminated! Saving the current progress...")
            #break
        if i>4:
            k = 4
        else:
            k = i
            
        gov_kills_tier1 = 0
        gov_kills_tier2 = 0
        gov_kills_tier3 = 0
        gov_kills_tier4 = 0
        gov_kills_tier5 = 0
        gov_victories = 0
        gov_defeats = 0
        gov_dead = 0
        gov_scouts = 0
        gov_rss_gathered = 0
        gov_rss_assistance = 0

        #Open governor
        device.shell(f'input tap 690 ' + str(Y[k]))
        time.sleep(1.5)
        
        ##### Ensure that governor tab is open #####
        gov_info = False
        count = 0
        while not (gov_info):
            image_check = device.screencap()
            with open(('check_more_info.png'), 'wb') as f:
                        f.write(image_check)
            image_check = cv2.imread('check_more_info.png',cv2.IMREAD_GRAYSCALE)
            roi = (313, 727, 137, 29)	
            im_check_more_info = image_check[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
            check_more_info = pytesseract.image_to_string(im_check_more_info,config="-c tessedit_char_whitelist=MoreInfo")
            if 'MoreInfo' not in check_more_info :
                device.shell(f'input swipe 690 605 690 540')
                device.shell(f'input tap 690 ' + str(Y[k]))
                count += 1
                time.sleep(2)
                if count == 5:
                    break
            else:
                gov_info = True
                break
        
        #nickname copy
        device.shell(f'input tap 690 283')
        time.sleep(0.5)
        
        ##### Governor main page capture #####
        image = device.screencap()
        with open(('gov_info.png'), 'wb') as f:
                    f.write(image)
        image = cv2.imread('gov_info.png')
        #Power and Killpoints
        roi = (770, 230, 200, 35)
        im_gov_id = image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        image = cv2.imread('gov_info.png')
        kernel = np.ones((2, 2), np.uint8)
     
        image = cv2.dilate(image, kernel) 
        roi = (898, 364, 180, 44)
        im_gov_power = image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1114, 364, 222, 44)
        im_gov_killpoints = image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        gov_name = tk.Tk().clipboard_get()
        roi = (645, 362, 260, 40) #alliance tag
        im_alliance_tag = image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        
        #kills tier
        device.shell(f'input tap 1118 350')
        
        #1st image OCR
        gov_id = read_ocr(im_gov_id)
        gov_power = read_ocr(im_gov_power)
        gov_killpoints = read_ocr(im_gov_killpoints)
        time.sleep(1)
        
        ##### Kill tier Capture #####
        image = device.screencap()
        with open(('kills_tier.png'), 'wb') as f:
                    f.write(image)
        image2 = cv2.imread('kills_tier.png') 	
        image2 = cv2.fastNlMeansDenoisingColored(image2,None,20,20,7,21) 
        roi = (863, 466, 215, 26) #tier 1
        im_kills_tier1 = image2[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]

        roi = (863, 511, 215, 26) #tier 2
        im_kills_tier2 = image2[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]

        roi = (863, 556, 215, 26) #tier 3
        im_kills_tier3 = image2[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]

        roi = (863, 601, 215, 26) #tier 4
        im_kills_tier4 = image2[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]

        roi = (863, 646, 215, 26) #tier 5
        im_kills_tier5 = image2[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]

        #More info tab
        device.shell(f'input tap 387 664') 
        
        ##### Kill tier OCR #####
        configed = '-c tessedit_char_whitelist=0123456789'
        gov_kills_tier1 = pytesseract.image_to_string(im_kills_tier1,config='-c page_separator='' '+configed)
        gov_kills_tier2 = pytesseract.image_to_string(im_kills_tier2,config='-c page_separator='' '+configed)
        gov_kills_tier3 = pytesseract.image_to_string(im_kills_tier3,config='-c page_separator='' '+configed)
        gov_kills_tier4 = pytesseract.image_to_string(im_kills_tier4,config='-c page_separator='' '+configed)
        gov_kills_tier5 = pytesseract.image_to_string(im_kills_tier5,config='-c page_separator='' '+configed)
        time.sleep(1)
        
        
        ##### More Info Page Capture #####
        image = device.screencap()
        with open(('more_info.png'), 'wb') as f:
                    f.write(image)
        image3 = cv2.imread('more_info.png')
        kernel = np.ones((2, 2), np.uint8)
        image3 = cv2.dilate(image3, kernel)
        roi = (1172 , 322 , 135 , 42) #Victories
        im_victories = image3[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1172 , 384 , 135 , 38) #Defeats
        im_defeats = image3[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130, 443, 183, 40) #dead
        im_dead = image3[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1172 , 505 , 135 , 43) #scouts
        im_scouts = image3[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130 , 608 , 183 , 45) #rss gathered
        im_rss_gathered = image3[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130, 668, 183, 40) #rss assistance
        im_rss_assistance = image3[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]

        
        #2nd check for deads with more filters to avoid some errors
        thresh = 127
        thresh_image = cv2.threshold(image3, thresh, 255, cv2.THRESH_BINARY)[1]
        roi = (1172 , 322 , 135 , 42) #Victories
        im_victories2 = thresh_image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1172 , 384 , 135 , 38) #Defeats
        im_defeats2 = roi = thresh_image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130, 443, 183, 40) #dead
        im_dead2 = thresh_image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1172 , 505 , 135 , 43) #scouts
        im_scouts2 = thresh_image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130 , 608 , 183 , 45) #rss gathered
        im_rss_gathered2 = thresh_image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130, 668, 183, 40) #rss assistance
        im_rss_assistance2 = thresh_image[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        
        #3rd check for deads with more filters to avoid some errors
        blur_img = cv2.GaussianBlur(image3, (3, 3), 0)
        roi = (1172 , 322 , 135 , 42) #Victories
        im_victories3 = blur_img[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1172 , 384 , 135 , 38) #Defeats
        im_defeats3 = blur_img[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130, 443, 183, 40) #dead
        im_dead3 = blur_img[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1172 , 505 , 135 , 43) #scouts
        im_scouts3 = roi = blur_img[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130 , 608 , 183 , 45) #rss gathered
        im_rss_gathered3 = blur_img[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]
        roi = (1130, 668, 183, 40) #rss assistance
        im_rss_assistance3 = blur_img[int(roi[1]):int(roi[1]+roi[3]), int(roi[0]):int(roi[0]+roi[2])]

        
        ##### More info page OCR #####
        gov_victories = read_ocr(im_victories)
        gov_victories2 = read_ocr(im_victories2)
        gov_victories3 = read_ocr(im_victories3)
        gov_defeats = read_ocr(im_defeats)
        gov_defeats2 = read_ocr(im_defeats2)
        gov_defeats3 = read_ocr(im_defeats3)
        gov_dead = read_ocr(im_dead)
        gov_dead2 = read_ocr(im_dead2)
        gov_dead3 = read_ocr(im_dead3)
        gov_scouts = read_ocr(im_scouts)
        gov_scouts2 = read_ocr(im_scouts2)
        gov_scouts3 = read_ocr(im_scouts3)
        gov_rss_gathered = read_ocr(im_rss_gathered)
        gov_rss_gathered2 = read_ocr(im_rss_gathered2)
        gov_rss_gathered3 = read_ocr(im_rss_gathered3)
        gov_rss_assistance = read_ocr(im_rss_assistance)
        gov_rss_assistance2 = read_ocr(im_rss_assistance2)
        gov_rss_assistance3 = read_ocr(im_rss_assistance3)
        
        
        ##### Alliance tag #####
        gray = cv2.cvtColor(im_alliance_tag,cv2.COLOR_BGR2GRAY)
        threshold_img = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
        alliance_tag = pytesseract.image_to_string(threshold_img,config='-c page_separator='' ')
        
        
        #Just to check the progress, printing in cmd the result for each governor
        if gov_power == '':
            gov_power = '0'
        if gov_killpoints =='':
            gov_killpoints = '0'
        if gov_kills_tier1 == '' :
            gov_kills_tier1 = '0'
        if gov_kills_tier2 == '' :
            gov_kills_tier2 = '0'
        if gov_kills_tier3 == '' :
            gov_kills_tier3 = '0'
        if gov_kills_tier4 == '' :
            gov_kills_tier4 = '0'
        if gov_kills_tier5 == '' :
            gov_kills_tier5 = '0'

        if gov_victories== '' : #victories
            if gov_victories2 == '' :
                if gov_victories3 == '' :
                    gov_victories = '0'
                else:
                    gov_victories = gov_victories3
            else:
                gov_victories = gov_victories2

        if gov_defeats == '' : #defeats
            if gov_defeats2 == '' :
                if gov_defeats3 == '' :
                    gov_defeats = '0'
                else :
                    gov_defeats = gov_defeats3
            else:
                gov_defeats = gov_defeats2

        if gov_dead == '' : #deads
            if gov_dead2 == '' :
                if gov_dead3 == '' :
                    gov_dead = '0'
                else :
                    gov_dead = gov_dead3
            else :
                gov_dead = gov_dead2

        if gov_scouts == '' : #scouts
            if gov_scouts2 == '' :
                if gov_scouts3 == '' :
                    gov_scouts = '0'
                else:
                    gov_scouts = gov_scouts3
            else:
                gov_scouts = gov_scouts2

        if gov_rss_gathered == '' : #rss_gathered
            if gov_rss_gathered2 == '' :
                if gov_rss_gathered3 == '' :
                    gov_rss_gathered = '0'
                else:
                    gov_rss_gathered = gov_rss_gathered3
            else:
                gov_rss_gathered = gov_rss_gathered2

        if gov_rss_assistance == '' : #rss_assistance
            if gov_rss_assistance == '' :
                if gov_rss_assistance3 == '' :
                    gov_rss_assistance = '0'
                else:
                    gov_rss_assistance = gov_rss_assistance3
            else:
                gov_rss_assistance = gov_rss_assistance2

        print(str(i+1)+'-'+'Governor Id: '+str(gov_id)+'\nName: '+str(gov_name)+'\nPower: '+str(gov_power)+'\nKillPoints: '+str(gov_killpoints)
        +'\nT1 kills: '+str(gov_kills_tier1)+'T2 kills: '+str(gov_kills_tier2)+'T3 kills: '+str(gov_kills_tier3)+'T4 kills: '+str(gov_kills_tier4)
        +'T5 kills: '+str(gov_kills_tier5)+'Victories: '+str(gov_victories)+'\nDefeats: '+str(gov_defeats)+'\nDeads: '+str(gov_dead)
        +'\nScouts: '+str(gov_scouts)+'\nRssGathered: '+str(gov_rss_gathered)+'\nRssAssistance: '+str(gov_rss_assistance)+'\n')

        device.shell(f'input tap 1396 58') #close more info
        time.sleep(0.5)
        device.shell(f'input tap 1365 104') #close governor info
        time.sleep(1)

        #Write results in excel file
        stats =[gov_id,gov_name,gov_power,gov_killpoints,gov_dead,
        gov_kills_tier1,gov_kills_tier2,gov_kills_tier3,
        gov_kills_tier4,gov_kills_tier5,gov_victories,
        gov_defeats,gov_scouts,gov_rss_gathered,gov_rss_assistance,alliance_tag]
        for f , b in enumerate(stats):
            sh.cell(row=i+2, column=f+1).value=b
        end_time = time.perf_counter()
        elapsed_time = end_time - start_time
        print("Elapsed time: ", elapsed_time)
except:
    print('An issue has occured. Please rerun the tool and use "resume scan option" from where tool stopped. If issue seems to remain, please contact me on discord!')
    #Save the excel file in the following format e.g. TOP300-2021-12-25-1253.xls or NEXT300-2021-12-25-1253.xls
    pass

if resume_scanning :
    file_name_prefix = 'NEXT'
else:
    file_name_prefix = 'TOP'
wb.save(file_name_prefix + str(search_range-j) + '-' +str(today)+ '-' + kingdom +'.xlsx')

#calculate totals
Top300power=0
Top600power=0
Top900power=0

for i in range(2,902):
    add=sh.cell(row=i,column=3).value
    if add == None:
        add=0
    if add > 150000000 :
        add = 150000000
    if i < 302 :
        Top300power += add
        Top600power += add
        Top900power += add
    if i < 602 and i > 301:
        Top600power += add
        Top900power += add
    if i < 902 and i > 601:
        Top900power += add


Top300kp=0
Top600kp=0
Top900kp=0

for i in range(2,902):
    add = sh.cell(row=i, column=4).value
    if add == None:
        add=0
    if i < 302 :
        Top300kp += add
        Top600kp += add
        Top900kp += add
    if i < 602 and i > 301:
        Top600kp += add
        Top900kp += add
    if i < 902 and i > 601:
        Top900kp += add

Top300T5=0
Top600T5=0
Top900T5=0

for i in range(2,902):
    add = sh.cell(row=i, column=10).value
    if add == None:
        add=0
    if i < 302 :
        Top300T5 += int(add)
        Top600T5 += int(add)
        Top900T5 += int(add)
    if i < 602 and i > 301:
        Top600T5 += int(add)
        Top900T5 += int(add)
    if i < 902 and i > 601:
        Top900T5 += int(add)

Top300T4=0
Top600T4=0
Top900T4=0

for i in range(2,902):
    add = sh.cell(row=i, column=9).value
    if add == None:
        add=0
    if i < 302 :
        Top300T4 += int(add)
        Top600T4 += int(add)
        Top900T4 += int(add)
    if i < 602 and i > 301:
        Top600T4 += int(add)
        Top900T4 += int(add)
    if i < 902 and i > 601:
        Top900T4 += int(add)

Top300deads=0
Top600deads=0
Top900deads=0

for i in range(2,902):
    add = sh.cell(row=i, column=5).value
    if add == None:
        add=0
    if i < 302 :
        Top300deads += add
        Top600deads += add
        Top900deads += add
    if i < 602 and i > 301:
        Top600deads += add
        Top900deads += add
    if i < 902 and i > 601:
        Top900deads += add

Top300rssgatherd=0
Top600rssgatherd=0
Top900rssgatherd=0

for i in range(2,902):
    add = sh.cell(row=i, column=14).value
    if add == None:
        add=0
    if i < 302 :
        Top300rssgatherd += add
        Top600rssgatherd += add
        Top900rssgatherd += add
    if i < 602 and i > 301:
        Top600rssgatherd += add
        Top900rssgatherd += add
    if i < 902 and i > 601:
        Top900rssgatherd += add

Top300rssassistance=0
Top600rssassistance=0
Top900rssassistance=0

for i in range(2,902):
    add = sh.cell(row=i, column=15).value
    if add == None:
        add=0
    if i < 302 :
        Top300rssassistance += add
        Top600rssassistance += add
        Top900rssassistance += add
    if i < 602 and i > 301 :
        Top600rssassistance += add
        Top900rssassistance += add
    if i < 902 and i > 601 :
        Top900rssassistance += add

#write totals
listed=[Top300power,Top600power,Top900power]
for i , b in enumerate(listed):
    sh.cell(row=2, column=i+19).value=b

listed=[Top300kp,Top600kp,Top900kp]
for i , b in enumerate(listed):
    sh.cell(row=3, column=i+19).value=b

listed=[Top300T5,Top600T5,Top900T5]
for i , b in enumerate(listed):
    sh.cell(row=4, column=i+19).value=b

listed=[Top300T4,Top600T4,Top900T4]
for i , b in enumerate(listed):
    sh.cell(row=5, column=i+19).value=b

listed=[Top300deads,Top600deads,Top900deads]
for i , b in enumerate(listed):
    sh.cell(row=6, column=i+19).value=b

listed=[Top300rssgatherd,Top600rssgatherd,Top900rssgatherd]
for i , b in enumerate(listed):
    sh.cell(row=7, column=i+19).value=b

listed=[Top300rssassistance,Top600rssassistance,Top900rssassistance]
for i , b in enumerate(listed):
    sh.cell(row=8, column=i+19).value=b


wb.save(file_name_prefix + str(search_range-j) + '-' +str(today)+ '-' + kingdom +'.xlsx')